# Name:         read_excel.PY
# Author:       dell
# Date:         2019/12/17
# Name:         read_excel
# Description: 从表格读取数据


from openpyxl import load_workbook



def read_excel(filename, sheet_name):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    # 最大列数
    col = ws.max_column + 1
    # 最大行数
    row = ws.max_row + 1
    test_data = ({ws.cell(1, j).value: ws.cell(i, j).value for j in range(1, col)} for i in range(2, row))
    return test_data


if __name__ == '__main__':
    from dir_config import testdatas_dir

    filename = testdatas_dir + "从模板创作.xlsx"
    test_datas = list(read_excel(filename, 'data'))
    print(test_datas)
